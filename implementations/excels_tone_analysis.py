from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
import time
import random
import json
from openai import OpenAI
import threading

client = OpenAI(
    api_key="sk-0ddd",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

prompt_guodong = """
        我会给你一句发生在一个在线医患平台患者发言的单条的对话，请根据如下需求分析
        1、将所有的发言文本进行【发言内容主题标签】分类，每个文本可以有多个分类标签（多个标签用","符号隔开），标签分类示例包括：
        用药相关咨询，示例：候大夫我每开哪几种药吃一粒。请问一下，前几天我买了肠胃的药，可以吃吗
        指标或数据异常咨询，示例：吴医生，打了甘精胰岛素10个单位，第一天血糖5.6，第二天5.5，今早上4.0了，怎么办？
        身体不适症状咨询，示例：您好我是病号曹茂岭每天晚上三点脑门疼出去怕光眼睛是怎么回事
        营养补品咨询，示例：李医生，多吃党参对血压有影响么。请问辅酶q10软胶囊这个保健品我可以吃吗
        运动锻炼咨询，示例：散步是热身运动吗？刚跑完8公里心率有点高
        医生门诊时间咨询，示例：主任早上什么时候有空。请问罗大夫本周四。是上午在门诊，还是下午在门诊？
        日常出行相关咨询，示例：我爸上周四安的支架，昨天才出院，本周五准备飞机出行，如果不能坐飞机可以坐动车不？
        感谢类发言，示例：今天早上我婆婆手术做了，很成功，没想费用这么少，谢谢大夫
        检查类咨询，示例：今天没有空腹  如果做检查影响吗
        指标测量相关，示例：服药前，醒后即测血压。
        复诊复查相关，示例：请问这个需要来医院复查吗
        饮食相关，示例：高血压能喝酒吗
        挂号相关，示例：你好，明天想李主任签字盖章，用不用挂号
        心脏相关，示例：心脏支架术后，心率要控制在多少之间？
        产品功能使用相关，示例：助力你好！可能是咱们的系统有问题，不能保存和提交。我进入这个悦压小程序了，没看到健康按钮啊？
        其他类。

        2、将所有的发言文本进行【语气分类】，分类示例包括：
        提问类，示例：心脏支架术后，心率要控制在多少之间？
        陈述类，示例：我住院化验葡萄糖5.37（MMOL）
        感叹类，示例：好的，谢谢！原来是不会这样啊！
        否定类，示例：不可能吧，所有抽血的结果都出来了，不可能就这一个结果不出来啊。
        其他类，以上都不是则归为其他类
        
        请返回如下json格式：
        {
            "发言内容主题标签":"{{你判断的发言主题标签，比如：用药相关咨询、指标或数据异常咨询、等等}}",
            "语气分类":"{{你判断的语气分类，比如：提问类、陈述类、感叹类、否定类、其他类}}",
            "reason":"{{你对上述两个判断的根据}}"
        }
        请注意：
        1. reason请简洁明了，给出你判断的理由，不超过50个字
        2. 发言内容主题标签和语气分类请用中文的tag，只在给出的tag中选择，不得自己发挥
        3. 返回结果按照以上要求，以json格式给出
    """


def call_LLM(text, prompt, max_retries=3, retry_delay=30):
    """调用大模型API，并处理速率限制错误"""
    
    for retry in range(max_retries):
        try:
            completion = client.chat.completions.create(
                model="qwen-plus",
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": text}
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
                # enable_thinking=False,
            )
            result = completion.choices[0].message.content
            usage = completion.usage
            print(f"\n $$$$$$$$$$ \n")
            
            print(f"请求token数（prompt）：{usage.prompt_tokens}")
            print(f"响应token数（completion）：{usage.completion_tokens}")
            print(f"总token数：{usage.total_tokens}")
            print(f"\n\n\n")
            
            return result
        except Exception as e:
            wait_time = retry_delay + random.uniform(0, 30)
            print(f"Error occurred: {e}. Retrying {retry} times in {wait_time} seconds...")
            time.sleep(wait_time)  # 随机延迟
    return '{"result": "未知", "reason": "重试次数超过最大值"}'

def is_medical_related(data):
    prompt = prompt_guodong
    rlt = call_LLM(data['content'], prompt)
    rlt_json = json.loads(rlt)
    print(f"{data['row']} 发言 {data['content']} 结果 {rlt_json}")
    content_tag = rlt_json["发言内容主题标签"]
    tone_tag = rlt_json["语气分类"]
    reason = rlt_json["reason"]
    return data['row'], content_tag, tone_tag, reason

def update_excel(row, content_tag, tone_tag, reason, sheet, lock):
    with lock:
        sheet.cell(row=row, column=6, value=content_tag)
        sheet.cell(row=row, column=7, value=tone_tag)
        sheet.cell(row=row, column=8, value=reason)

def process(min, max):
    
    start = time.time()
    # 打开Excel文件
    workbook = openpyxl.load_workbook('/Users/m677418/Downloads/dd.xlsx')
    
    content_col = 4
    
    # 遍历每个工作表
    for sheet_name in ['群聊', '1v1']:
        sheet = workbook[sheet_name]
        count = 0
        # 预处理所有需要处理的数据
        cell_list = []
        for row in sheet.iter_rows(min_row=min, max_row=max, min_col=1, max_col=8, values_only=False):
            # print(f"row: {row[0].row} content: {row[content_col].value} row[5]: {row[5].value} len len(row[content_col].value): {len(row[content_col].value)}")
            # print(f"Not row[5]: {not row[5]} type(row[5].value): {type(row[5].value)} row[5]: {row[5].value}")
            if row[content_col].value and (not row[5].value) and len(row[content_col].value) > 5:
                # print(f"appending celllist ")
                cell_list.append({
                    "row": row[0].row,
                    "content": row[content_col].value
                })
        
        # 创建一个锁
        lock = threading.Lock()
        
        # 多线程处理
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = []
            for cells in cell_list:
                future = executor.submit(is_medical_related, cells)
                futures.append(future)
            
            for future in as_completed(futures):
                row, content_tag, tone_tag, reason = future.result()
                print(f'Process {row}: {content_tag} {tone_tag} {reason}')
                update_excel(row, content_tag, tone_tag, reason, sheet, lock)
    
    end = time.time()
    process_time = int(end - start)
    
    # 保存更新后的Excel文件
    workbook.save(f"/Users/m677418/Downloads/ddt.xlsx")

if __name__ == "__main__":
    start = time.time()
    process(2, 12400)
    end = time.time()
    duration_in_seconds = int(end - start)
    print(f"process time: {duration_in_seconds} seconds")
