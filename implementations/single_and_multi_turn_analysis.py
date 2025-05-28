from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
import time
import random
import json
from openai import OpenAI
import threading
import pandas as pd
from functools import lru_cache

client = OpenAI(
    api_key="sk-dd",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# 全局常量定义
CONTENT_COL = 8
CHAT_ID_COL = 4
ROLE_COL = 2
SENDER_COL = 3
RESULT_COL = 12
REASON_COL = 13

# 预加载并缓存提示词
@lru_cache(maxsize=1)
def get_prompt():
    return  """
          

        ### **打标签说明**  
        - **最优匹配**：请给回话匹配一个最接近的标签，只有一个标签。  
        - **场景区分**：  
        - “治疗/用药方案咨询反馈” 仅包含询问用药方式（如频次、时间），不涉及剂量调整；  
        - **角色关系**：平台中有医生、患者、、流程引导）。
        请严格按如下json格式返回结果：
        {
            "result":"{{科普答疑 或者 加减药量 等你判断出来的唯一标签}}",
            "reason":"{{你判断的根据}}"
        }
        请注意：
        1. reason请简洁明了，给出你判断的理由，不超过20个字
        2. result请用中文的标签，只在给出的标签中选择，不得自己发挥
    """

def call_LLM(text, max_retries=3, retry_delay=30):
    """调用大模型API，并处理速率限制错误"""
    prompt = get_prompt()
    for retry in range(max_retries):
        try:
            completion = client.chat.completions.create(
                model="qwen-plus",
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": text}
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            result = completion.choices[0].message.content
            usage = completion.usage
            print(f"\n $$$$$$$$$$ \n")
            print(f"请求token数（prompt）：{usage.prompt_tokens}")
            print(f"响应token数（completion）：{usage.completion_tokens}")
            print(f"总token数：{usage.total_tokens}")
            print(f"\n\n\n")
            return result
        except Exception as e:
            wait_time = retry_delay + random.uniform(0, 30)
            print(f"Error occurred: {e}. Retrying {retry} times in {wait_time} seconds...")
            time.sleep(wait_time)  # 随机延迟
    return '{"result": "未知", "reason": "重试次数超过最大值"}'


def is_medical_related(row_data, all_data):
    """处理医生发言（单条打标签）"""
    row_id = row_data['row']
    content = row_data['content']
    chat_id = row_data['chat_id']
    
    # 单轮对话处理（仅医生发言打标签）
    rlt = call_LLM(content)
    rlt_json = json.loads(rlt)
    print(f"{row_id} 医生发言 {content} 结果 {rlt_json}")
    result = rlt_json["result"]
    reason = rlt_json["reason"]
    
    # 多轮对话处理（回溯所有角色消息）
    if result == "确认性回复":
        result, reason = multi_turn_processing(row_id, chat_id, all_data)
    
    return row_id, result, reason

def multi_turn_processing(row_id, chat_id, all_data, k=5):
    """回溯所有角色的消息（医生/患者/小悦）"""
    print(f"多轮对话处理: row={row_id}, chat_id={chat_id}, k={k} \n")
    conversations = []
    
    # 逆序遍历所有行（从当前行向上查找，包含所有角色）
    for current_row in range(row_id + 1, 0, -1):
        if current_row not in all_data:
            continue
        
        msg = all_data[current_row]
        # 仅保留有内容且chat_id匹配的消息（不限角色）
        if msg['content'] and msg['chat_id'] == chat_id:
            conversations.append({
                "sender": msg['sender'],  # 
                "content": msg['content'],
                "role": msg['role']       # 记录角色用于上下文分析
            })
            if len(conversations) >= k:
                break
    
    # 按时间顺序排列（旧→新）
    conversations.reverse()
    
    if conversations:
        # 传递角色信息给LLM，增强上下文理解
        formatted_conv = json.dumps(conversations, ensure_ascii=False)
        rlt = call_LLM(formatted_conv)
        print(f"多轮对话内容: row = {row_id} \n {conversations} \n")
        print(f"多轮对话判断结果: {rlt} \n")
        rlt_json = json.loads(rlt)
        return rlt_json["result"] + '_多轮标签', rlt_json["reason"]
    
    return "确认性回复_多轮标签", "无有效历史对话"

def process(min_row, max_row):
    start = time.time()
    workbook = openpyxl.load_workbook('/Users/m677418/Downloads/20250331社群会话记录.xlsx')
    sheet = workbook['chat']
    
    # 预加载所有行数据（包括所有角色）
    all_data = {}  # 键：行号，值：{chat_id, content, role, sender}
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
        row_num = row[0].row
        content = row[CONTENT_COL].value
        if not content:
            continue  # 跳过内容为空的行（无论角色）
        
        all_data[row_num] = {
            'row': row_num,
            'chat_id': row[CHAT_ID_COL].value,
            'content': content,
            'role': row[ROLE_COL].value,    # 新增：记录角色
            'sender': row[SENDER_COL].value
        }
    
    # 筛选需要处理的医生发言（单条打标签仅处理医生）
    doctor_rows = [row_data for row_data in all_data.values() if row_data['role'] == "医生"]
    
    lock = threading.Lock()
    
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for row_data in doctor_rows:
            future = executor.submit(is_medical_related, row_data, all_data)
            futures.append(future)
        
        for future in as_completed(futures):
            try:
                row, result, reason = future.result()
                with lock:
                    sheet.cell(row=row, column=RESULT_COL, value=result)
                    sheet.cell(row=row, column=REASON_COL, value=reason)
            except Exception as e:
                print(f"行 {row} 处理失败: {e}")
    
    end = time.time()
    print(f"处理时间: {int(end - start)}秒, 共处理医生发言: {len(doctor_rows)}")
    workbook.save(f"/Users/m677418/Downloads/20250331社群会话记录_rlt_v4.xlsx")

if __name__ == "__main__":
    process(2, 5300)
