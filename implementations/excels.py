from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
import time
import random
import json
from openai import OpenAI
import threading

client = OpenAI(
    api_key="sk-ea5",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

prompt_guodong = """
        我会给你一句单条的对话，请判断它是否是有意义的跟医疗有关的信息。比如：“谢谢”， "好的"，"好"，就该返回"否"，
        "吃药前家里的", "膝盖疼"，“走路多”，“血压”等则是跟医疗医药相关的，请返回"是"
        如果是则返回是，并给出你的理由，如果不是则返回否，
        同样给出你的理由。
        输出请用如下json格式：
        {
            "result":"是 或者 否",
            "reason":"{{你判断的根据}}"
        }
    """
prompt_zhou_doc = """
    我会给你一句医生给患者的对话，请判断它是否跟医生想让患者回到医院有关，比如让患者回医院复诊，做手术，回医院检查等等。
    请返回是或者否，同样给出你的理由。
    输出请用如下json格式：
    {
        "result":"医生发言_是 或者 医生发言_否",
        "reason":"{{你判断的根据}}"
    }
"""

prompt_zhou_patient = """
    我会给你一句单条的对话，是来自患者，说话的对象是医生。请判断它是否是在像医生求助和做医疗方面的咨询和信息交互。如果是则返回"是"，如果不是则返回"否"。
    请记住，一些寒暄、客套、感谢、等都应该返回"否"
    同同时给出你的理由。
    输出请用如下json格式：
    {
        "result":"患者发言_是 或者 患者发言_否",
        "reason":"{{你判断的根据}}"
    }
    例如： "周主任.不好意思打扰你啦。[强][强]","@医生助理小悦 知道了，谢谢" 等都该返回 否。
"""

def call_LLM(text, prompt, max_retries=3, retry_delay=30):
    """调用大模型API，并处理速率限制错误"""
    
    for retry in range(max_retries):
        try:
            completion = client.chat.completions.create(
                model="qwen-plus",
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": text}
                ],
                temperature=0.3,
                response_format={"type": "json_object"}
            )
            result = completion.choices[0].message.content
            return result
        except Exception as e:
            wait_time = retry_delay + random.uniform(0, 30)
            print(f"Error occurred: {e}. Retrying {retry} times in {wait_time} seconds...")
            time.sleep(wait_time)  # 随机延迟
    return '{"result": "未知", "reason": "重试次数超过最大值"}'

def is_medical_related(cells):
    prompt = ""
    if cells[0].value == "医生":
        prompt = prompt_zhou_doc
    else:
        prompt = prompt_zhou_patient
    rlt = call_LLM(cells[1].value, prompt)
    rlt_json = json.loads(rlt)
    print(f"{cells[0].value} 发言 {cells[1].value} 结果 {rlt_json}")
    result = rlt_json["result"]
    reason = rlt_json["reason"]
    return cells[0].row, result, reason

def update_excel(row, result, reason, sheet, lock):
    with lock:
        sheet.cell(row=row, column=15, value=result)
        sheet.cell(row=row, column=16, value=reason)

def process(min, max):
    # 打开Excel文件
    workbook = openpyxl.load_workbook('/Users/m677418/projects/csvs/tobe_tagged.xlsx')
    
    # 遍历每个工作表
    for sheet_name in ['周发展']:
        sheet = workbook[sheet_name]
        count = 0
        # 预处理所有需要处理的数据
        cell_list = []
        for row in sheet.iter_rows(min_row=min, max_row=max, min_col=1, max_col=10, values_only=False):
            cell1 = row[0]
            cell2 = row[9]
            cells = [cell1, cell2]
            if cell1.value and cell2.value:  # 如果J列数据不为空
                cell_list.append(cells)
        
        # 创建一个锁
        lock = threading.Lock()
        
        # 多线程处理
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = []
            for cells in cell_list:
                future = executor.submit(is_medical_related, cells)
                futures.append(future)
            
            for future in as_completed(futures):
                row, result, reason = future.result()
                print(f'Process {row}:')
                update_excel(row, result, reason, sheet, lock)
    
    # 保存更新后的Excel文件
    workbook.save(f"zhou_{min}_{max}.xlsx")

if __name__ == "__main__":
    process(1001, 3000)

